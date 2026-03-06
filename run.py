#!/usr/bin/env python3
"""CLI wrapper for the DuckDuckGo job board crawler."""
import argparse
import json
import time
from pathlib import Path
from typing import List, Dict, Optional

from crawler import DuckDuckGoJobBoardCrawler
import requests
from bs4 import BeautifulSoup
import re
import os
from datetime import datetime
try:
    from supabase_mcp import has_config, insert_discoveries
except Exception:
    has_config = lambda: False
    insert_discoveries = lambda rows: False


def parse_args():
    p = argparse.ArgumentParser(description="DuckDuckGo job board crawler with resume support")
    p.add_argument("--queries", nargs="+", help="Search queries to run", required=False)
    p.add_argument("--pages", type=int, default=1, help="Pages per query")
    p.add_argument("--rate", type=float, default=1.0, help="Seconds between requests")
    p.add_argument("--output", default="results.json", help="Base output path (auto-suffixed)")
    p.add_argument(
        "--engine",
        choices=["duckduckgo", "brave", "playwright"],
        default="brave",
        help="Which search engine backend to use",
    )
    p.add_argument(
        "--filter",
        action="store_true",
        help="Keep only URLs that match job-board heuristics",
    )
    p.add_argument(
        "--verify",
        action="store_true",
        help="When filtering, fetch each homepage and scan for job keywords",
    )
    p.add_argument(
        "--discover",
        type=int,
        default=None,
        help="Stop when this many NEW job boards are discovered in this run",
    )
    p.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Alias for --discover (backward compatible)",
    )
    p.add_argument(
        "--detect-careers",
        action="store_true",
        help="Attempt to find career pages for company domains",
    )
    p.add_argument(
        "--resume",
        action="store_true",
        help="Resume from last checkpoint if available",
    )
    p.add_argument(
        "--checkpoint",
        default=".crawler_checkpoint.json",
        help="Checkpoint file path",
    )
    p.add_argument(
        "--chambers",
        action="store_true",
        help="Run Chamber-of-Commerce directory discovery (state -> chamber -> members)",
    )
    p.add_argument(
        "--states",
        nargs="+",
        help="List of states to search (e.g. 'California Texas'). If omitted, uses all US states",
    )
    p.add_argument(
        "--state",
        help="Run for a single state (overrides --states)",
    )
    p.add_argument(
        "--chamber-output",
        default="chambers.xlsx",
        help="Output workbook for chamber/member scraping",
    )
    p.add_argument(
        "--proxies",
        help="Path to a newline-separated file of HTTP/HTTPS proxies (e.g. http://user:pass@host:port)",
    )
    p.add_argument(
        "--proxy-ban-seconds",
        type=int,
        default=300,
        help="Seconds to ban a proxy after failure",
    )
    p.add_argument(
        "--max-retries",
        type=int,
        default=3,
        help="Max retries for requests",
    )
    p.add_argument(
        "--backoff-factor",
        type=float,
        default=1.0,
        help="Backoff factor for retries",
    )
    p.add_argument(
        "--jitter",
        type=float,
        default=0.2,
        help="Jitter multiplier for backoff sleeps",
    )
        default="chambers.xlsx",
        help="Output workbook for chamber/member scraping",
    )
    return p.parse_args()


def categorize_urls(crawler, urls: List[str]) -> Dict[str, List[Dict]]:
    """Categorize URLs into aggregators and company career pages."""
    aggregators = []
    companies = []
    
    for url in urls:
        domain = crawler.get_domain(url)
        if not domain:
            continue
        
        if crawler.is_job_aggregator(url):
            aggregators.append({
                "url": url,
                "domain": domain,
                "title": domain,
            })
        else:
            companies.append({
                "url": url,
                "domain": domain,
                "title": domain,
                "career_page": url,
            })
    
    return {"aggregators": aggregators, "companies": companies}


# canonical list of US states (50) — we'll sort before use to ensure alphabetical order
ALL_US_STATES = [
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi",
    "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey",
    "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio", "Oklahoma",
    "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
]


def get_all_states() -> List[str]:
    """Return all 50 US states in alphabetical order."""
    return sorted(ALL_US_STATES)


def fetch_and_save_counties(wiki_url: str, output_path: str = "us_counties.xlsx") -> None:
    """Fetch county list from Wikipedia and save to an Excel workbook.

    This scans 'wikitable' elements and attempts to extract county/state columns,
    falling back to the first two columns when headers aren't explicit.
    """
    try:
        resp = requests.get(wiki_url, timeout=20, headers={"User-Agent": "jobboard-crawler/1.0"})
        resp.raise_for_status()
    except Exception as e:
        print(f"  Could not fetch county data: {e}")
        return

    soup = BeautifulSoup(resp.text, "html.parser")
    tables = soup.find_all("table", class_=lambda c: c and "wikitable" in c)
    rows = []  # list of (state, county)

    for table in tables:
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        county_idx = None
        state_idx = None
        for i, h in enumerate(headers):
            h_low = h.lower()
            if "county" in h_low or "county or equivalent" in h_low or h_low.startswith("name"):
                county_idx = i
            if "state" in h_low or "state or equivalent" in h_low:
                state_idx = i

        if county_idx is None or state_idx is None:
            county_idx = 0
            state_idx = 1

        for tr in table.find_all("tr")[1:]:
            cols = tr.find_all(["td", "th"])
            if len(cols) <= max(county_idx, state_idx):
                continue
            county = cols[county_idx].get_text(" ", strip=True)
            state = cols[state_idx].get_text(" ", strip=True)
            county = re.sub(r"\[.*?\]", "", county).strip()
            state = re.sub(r"\[.*?\]", "", state).strip()
            if state and county:
                rows.append((state, county))

    if not rows:
        print("  No county rows extracted from Wikipedia page")
        return

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Counties"
        ws.append(["State", "County"])
        for s, c in rows:
            ws.append([s, c])
        wb.save(output_path)
        print(f"  Saved {len(rows)} county rows to {output_path}")
    except Exception as e:
        print(f"  Error saving county workbook: {e}")


def augment_counties_with_chambers(crawler, counties_path: str = "us_counties.xlsx", output_path: Optional[str] = None, max_search_pages: int = 1):
    """Read the counties workbook and search each county for its chamber directory.

    Adds a "Chamber Directory" column to the workbook and writes the discovered
    directory URL (first match) for each county. If `output_path` is None the
    input file is updated in-place.
    """
    out_path = output_path or counties_path
    try:
        from openpyxl import load_workbook
        wb = load_workbook(counties_path)
        ws = wb.active
    except Exception as e:
        print(f"  Could not open counties workbook: {e}")
        return

    # ensure header
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "Chamber Directory" not in header:
        ws.cell(row=1, column=len(header) + 1, value="Chamber Directory")
        header.append("Chamber Directory")

    state_col = None
    county_col = None
    for idx, h in enumerate(header, start=1):
        if h and h.lower().strip() == "state":
            state_col = idx
        if h and h.lower().strip() == "county":
            county_col = idx

    if state_col is None or county_col is None:
        print("  Counties workbook missing expected 'State'/'County' columns")
        return

    total = max(0, ws.max_row - 1)
    found = 0
    # determine chamber column index
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        chamber_col = header.index("Chamber Directory") + 1
    except ValueError:
        chamber_col = len(header) + 1

    for i, row_idx in enumerate(range(2, ws.max_row + 1), start=1):
        state = (ws.cell(row=row_idx, column=state_col).value or "").strip()
        county = (ws.cell(row=row_idx, column=county_col).value or "").strip()
        if not county or not state:
            continue

        # build query: e.g. 'Uinta wyoming chamber of commerce business directory'
        query = f"{county} {state} chamber of commerce business directory"
        # attempt search with retries and exponential backoff on 429s
        results = []
        max_retries = 5
        backoff_base = 2
        tried_alt = False
        for attempt in range(1, max_retries + 1):
            try:
                results = crawler.search(query, pages=max_search_pages)
                break
            except Exception as e:
                # detect HTTP 429 if possible
                from requests import exceptions as req_ex
                is_429 = False
                if isinstance(e, req_ex.HTTPError) and getattr(e, 'response', None) is not None:
                    try:
                        if e.response.status_code == 429:
                            is_429 = True
                    except Exception:
                        pass
                # fall back to string check if type not available
                if not is_429 and isinstance(e, Exception) and '429' in str(e):
                    is_429 = True

                if is_429:
                    wait = backoff_base ** attempt
                    print(f"    request failed: 429 Too Many Requests; backing off {wait}s (attempt {attempt}/{max_retries})")
                    time.sleep(wait)
                    # if repeated 429s and we haven't tried an alternate engine, try DuckDuckGo once
                    if attempt >= 3 and not tried_alt:
                        try:
                            print("    trying alternate search engine (duckduckgo) due to repeated 429s")
                            alt = DuckDuckGoJobBoardCrawler(
                                rate_limit=max(crawler.rate_limit, 2.0),
                                engine="duckduckgo",
                                max_retries=crawler.max_retries,
                                backoff_factor=crawler.backoff_factor,
                                jitter=crawler.jitter,
                                proxies=getattr(crawler, 'proxies', None),
                                proxy_ban_seconds=getattr(crawler, 'proxy_ban_seconds', 300),
                            )
                            results = alt.search(query, pages=max_search_pages)
                            tried_alt = True
                            break
                        except Exception as ae:
                            print(f"    alternate engine attempt failed: {ae}")
                            # continue retry/backoff loop
                            continue
                    continue
                else:
                    print(f"    Search error for {county}, {state}: {e}")
                    break

        chamber_url = ""
        for r in results:
            u = r.get("url")
            if not u:
                continue
            d = crawler.get_domain(u)
            if not d:
                continue
            if "chamber" in d or "chamber" in u.lower():
                chamber_url = u
                break

        # write to the Chamber Directory column and save immediately when found
        ws.cell(row=row_idx, column=chamber_col, value=chamber_url)
        if chamber_url:
            found += 1
            try:
                wb.save(out_path)
            except Exception as e:
                print(f"  Error saving augmented counties workbook (row {row_idx}): {e}")

        # progress print
        print(f"    [{i}/{total}] {county}, {state} -> {chamber_url or '---'}")

        # be polite
        time.sleep(crawler.rate_limit)

    print(f"  Augmented counties workbook progress complete ({found}/{total} chambers found)")


def main():
    args = parse_args()
    discover_target = args.discover if args.discover is not None else args.limit
    if discover_target is not None and discover_target <= 0:
        raise SystemExit("--discover/--limit must be greater than 0")
    
    # load checkpoint if resuming
    checkpoint = DuckDuckGoJobBoardCrawler.load_checkpoint(args.checkpoint)
    
    queries: List[str] = args.queries or DuckDuckGoJobBoardCrawler.default_queries()
    
    # resume from last query
    start_index = 0
    if args.resume:
        start_index = checkpoint.get("last_query_index", -1) + 1
        if start_index > 0:
            print(f"[Resume] Starting from query #{start_index} (discovered: {checkpoint.get('discovered_count', 0)})")
    
    # load proxies file if provided
    proxies_list = None
    if args.proxies:
        try:
            ppath = Path(args.proxies)
            if ppath.exists():
                with open(ppath, "r", encoding="utf-8") as pf:
                    lines = [l.strip() for l in pf.readlines()]
                    # ignore blank lines and comments
                    proxies_list = [l for l in lines if l and not l.startswith("#")]
        except Exception as e:
            print(f"Warning: could not read proxies file: {e}")

    crawler = DuckDuckGoJobBoardCrawler(
        rate_limit=args.rate,
        engine=args.engine,
        max_retries=args.max_retries,
        backoff_factor=args.backoff_factor,
        jitter=args.jitter,
        proxies=proxies_list,
        proxy_ban_seconds=args.proxy_ban_seconds,
    )
    seen_keys = set(checkpoint.get("discovered_keys", [])) if args.resume else set()
    if args.resume and not seen_keys:
        previous_count = checkpoint.get("discovered_count", 0)
        if previous_count:
            print(f"[Resume] Previous discovered count found ({previous_count}) but no key history; counting new discoveries from this run.")
    newly_discovered_urls: List[str] = []
    
    # Chambers workflow
    if args.chambers:
        print("Running Chamber-of-Commerce directory discovery...")
        # determine which states to run: single-state flag overrides --states
        if args.state:
            states = [args.state]
        else:
            states = args.states or get_all_states()

        # fetch and save US counties from Wikipedia before proceeding with chamber discovery
        print("  Fetching US counties from Wikipedia and saving to 'us_counties.xlsx'...")
        try:
            fetch_and_save_counties(
                "https://en.wikipedia.org/wiki/List_of_United_States_counties_and_county_equivalents",
                output_path="us_counties.xlsx",
            )
        except Exception as e:
            print(f"  County extraction failed: {e}")

        # Augment the counties workbook with discovered chamber directory links
        print("  Augmenting counties workbook with chamber directory links...")
        try:
            augment_counties_with_chambers(crawler, counties_path="us_counties.xlsx")
        except Exception as e:
            print(f"  County->chamber augmentation failed: {e}")

        chamber_progress = checkpoint.get("chamber_progress", {})
        members_rows = []
        verified_rows = []

        total_states = len(states)
        for si, state in enumerate(states, start=1):
            print(f"[{si}/{total_states}] State: {state}")
            query = f"{state} chamber of commerce member directory"
            try:
                results = crawler.search(query, pages=1)
            except Exception as e:
                print(f"  Search error for {state}: {e}")
                continue

            chamber_sites = []
            for r in results:
                u = r.get("url")
                if not u:
                    continue
                d = crawler.get_domain(u)
                if not d:
                    continue
                if "chamber" in d or "chamber" in u.lower():
                    chamber_sites.append(u)

            chamber_sites = crawler.dedupe(chamber_sites)[:10]

            for site in chamber_sites:
                if chamber_progress.get(site):
                    print(f"  Skipping already-processed chamber: {site}")
                    continue
                print(f"  Inspecting chamber site: {site}")
                directory = crawler.find_member_directory(site)
                if not directory:
                    print("    No member directory found")
                    chamber_progress[site] = True
                    checkpoint["chamber_progress"] = chamber_progress
                    DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)
                    continue

                print(f"    Found directory: {directory} — scraping members...")
                members = crawler.scrape_member_directory(directory, max_pages=50)
                print(f"    Extracted {len(members)} members")

                for m in members:
                    name = m.get("name")
                    url = m.get("url")
                    members_rows.append((state, site, name, url))

                    # quick company site check
                    if not crawler.confirm_company_site(url, deep=bool(args.verify)):
                        continue

                    # try to find career links and (optionally) verify
                    career_candidates = crawler.find_career_links(url)
                    found_career = None
                    if career_candidates:
                        if bool(args.verify):
                            for cand in career_candidates:
                                if crawler.verify_jobs_on_page(cand):
                                    found_career = cand
                                    break
                        else:
                            # quick mode: record the first candidate without deep fetch
                            found_career = career_candidates[0]

                    if found_career:
                        domain = crawler.get_domain(url)
                        verified_rows.append((state, site, name, url, domain, found_career))
                        # register discovery key to avoid duplicates later
                        seen_keys.add(domain or url)

                # mark chamber processed and save checkpoint
                chamber_progress[site] = True
                checkpoint["chamber_progress"] = chamber_progress
                checkpoint["discovered_keys"] = list(seen_keys)
                checkpoint["discovered_count"] = len(seen_keys)
                DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)

        # save to workbook
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Members"
            ws1.append(["State", "Chamber Site", "Member Name", "Member URL"])
            for r in members_rows:
                ws1.append(list(r))

            ws2 = wb.create_sheet("Verified Careers")
            ws2.append(["State", "Chamber Site", "Company Name", "Company URL", "Domain", "Career Page"])
            for r in verified_rows:
                ws2.append(list(r))

            wb.save(args.chamber_output)
            print(f"Saved chamber scraping results to {args.chamber_output}")
        except Exception as e:
            print(f"Error saving chamber workbook: {e}")

        # update checkpoint
        checkpoint["last_query_index"] = checkpoint.get("last_query_index", -1)
        checkpoint["discovered_keys"] = list(seen_keys)
        checkpoint["discovered_count"] = len(seen_keys)
        DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)

        print("Chamber-of-Commerce run complete.")
        return
    # run queries
    for idx, q in enumerate(queries[start_index:], start=start_index):
        if discover_target and len(newly_discovered_urls) >= discover_target:
            print(f"[Target reached] Discovered {len(newly_discovered_urls)} new job boards")
            break
        
        print(f"[{idx + 1}/{len(queries)}] Searching: {q}")
        try:
            items = crawler.search(q, pages=args.pages)
            urls = crawler.extract_links(items)
            urls = crawler.dedupe(urls)

            if args.filter:
                urls = crawler.filter_urls(urls, verify_content=args.verify)

            new_for_query = 0
            for url in urls:
                domain = crawler.get_domain(url)
                unique_key = domain or url
                if unique_key in seen_keys:
                    continue
                seen_keys.add(unique_key)
                newly_discovered_urls.append(url)
                new_for_query += 1
                if discover_target and len(newly_discovered_urls) >= discover_target:
                    break

            print(f"  New discovered in query: {new_for_query} (total new: {len(newly_discovered_urls)})")
            
            # save checkpoint
            checkpoint["last_query_index"] = idx
            checkpoint["discovered_count"] = len(seen_keys)
            checkpoint["discovered_keys"] = list(seen_keys)
            DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)

            if discover_target and len(newly_discovered_urls) >= discover_target:
                print(f"[Target reached] Discovered {len(newly_discovered_urls)} new job boards")
                break
        except KeyboardInterrupt:
            print("\n[Checkpoint saved] Run with --resume to continue from here")
            checkpoint["last_query_index"] = idx - 1
            checkpoint["discovered_count"] = len(seen_keys)
            checkpoint["discovered_keys"] = list(seen_keys)
            DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)
            return
        except Exception as e:
            print(f"  Error: {e}")

    urls = newly_discovered_urls
    print(f"Discovered {len(urls)} new unique job board URLs in this run")
    
    # categorize
    print("Categorizing URLs...")
    categorized = categorize_urls(crawler, urls)
    aggregators = categorized["aggregators"]
    companies = categorized["companies"]
    
    print(f"  Aggregators: {len(aggregators)}")
    print(f"  Company career pages (candidates): {len(companies)}")

    # First ensure non-aggregator candidates are actually company websites.
    # Use deep check when --verify is requested.
    company_candidates = []
    for c in companies:
        domain = c.get("domain")
        check_url = c.get("url")
        is_company = crawler.confirm_company_site(check_url, deep=bool(args.verify))
        if not is_company:
            print(f"  Skipping non-company domain: {domain} ({check_url})")
            continue
        company_candidates.append(c)

    # detect career pages for the filtered company candidates if requested
    if args.detect_careers:
        print("Detecting career pages...")
        for company in company_candidates:
            domain = company["domain"]
            print(f"  Checking {domain} for career page...")
            career_page = crawler.find_career_page(domain)
            if career_page:
                company["career_page"] = career_page
                print(f"    Found: {career_page}")

    # Validate that company rows are actual job boards. If --verify was
    # requested we perform a deeper scrape; otherwise use lightweight check.
    validated_companies = []
    for c in company_candidates:
        check_url = c.get("career_page") or c.get("url")
        is_board = crawler.confirm_job_board(check_url, deep=bool(args.verify))
        if is_board:
            validated_companies.append(c)
        else:
            print(f"  Skipping non-job-board: {c.get('domain')} ({check_url})")
    companies = validated_companies
    
    # determine output paths
    base_output = Path(args.output)
    if base_output.suffix:
        base_name = str(base_output.parent / base_output.stem)
    else:
        base_name = str(base_output)
    
    company_file = f"{base_name}_companies.xlsx"
    aggregator_file = f"{base_name}_aggregators.xlsx"
    
    # save to spreadsheets (always exactly two output files)
    crawler.save_company_careers_to_excel(companies, company_file)
    print(f"Saved {len(companies)} company career pages to {company_file}")

    crawler.save_aggregators_to_excel(aggregators, aggregator_file)
    print(f"Saved {len(aggregators)} job aggregators to {aggregator_file}")
    
    # optionally push discoveries to Supabase if configured
    try:
        if has_config() and companies:
            rows = []
            for c in companies:
                rows.append({
                    "domain": c.get("domain"),
                    "url": c.get("career_page") or c.get("url"),
                    "source": c.get("source", "crawler"),
                    "discovered_at": datetime.utcnow().isoformat(),
                })
            insert_discoveries(rows)
    except Exception as e:
        print(f"  Supabase push error: {e}")

    # update checkpoint
    checkpoint["last_query_index"] = len(queries) - 1
    checkpoint["discovered_count"] = len(seen_keys)
    checkpoint["discovered_keys"] = list(seen_keys)
    DuckDuckGoJobBoardCrawler.save_checkpoint(checkpoint, args.checkpoint)

    print(f"\n✓ Crawl complete! ({len(urls)} new job boards discovered in this run)")


if __name__ == "__main__":
    main()
