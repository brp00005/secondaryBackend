"""
Scraper for US Chamber of Commerce directories
Scrapes chamber listings from https://www.uschamber.com/co/chambers/{state}
"""

import argparse
import pandas as pd
from bs4 import BeautifulSoup
import sys
import os
import re
import json
import time
from typing import List, Dict, Optional
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote

# Use tqdm if available for progress bars
try:
    from tqdm.auto import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    tqdm = None  # type: ignore
    TQDM_AVAILABLE = False

# Ensure the current directory and parent are in the path
sys.path.insert(0, os.getcwd())

try:
    from scripts.helpers import (
        fetch_with_retries,
        normalize_site,
        dedupe_key,
        write_workbook,
        HEADERS
    )
except ImportError:
    # If that fails, try to import from local script folders (including legacy locations)
    sys.path.insert(0, os.path.join(os.getcwd(), 'scripts'))
    sys.path.insert(0, os.path.join(os.getcwd(), 'old code v2'))
    from helpers import (
        fetch_with_retries,
        normalize_site,
        dedupe_key,
        write_workbook,
        HEADERS
    )

# Base URL for US Chamber of Commerce
USCHAMBER_BASE_URL = "https://www.uschamber.com/co/chambers"

# Output columns (in preferred order)
OUTPUT_COLUMNS = [
    'chamber_name',
    'city',
    'state',
    'website',
    'website_guess',
    'website_source',
    'maps_link',
    'member_directory_url',
    'all_links',
    'phone',
    'email',
    'address',
    'image',
]

# State slug mappings (state_name -> URL slug)
STATE_SLUGS = {
    'alabama': 'alabama',
    'alaska': 'alaska',
    'arizona': 'arizona',
    'arkansas': 'arkansas',
    'california': 'california',
    'colorado': 'colorado',
    'connecticut': 'connecticut',
    'delaware': 'delaware',
    'florida': 'florida',
    'georgia': 'georgia',
    'hawaii': 'hawaii',
    'idaho': 'idaho',
    'illinois': 'illinois',
    'indiana': 'indiana',
    'iowa': 'iowa',
    'kansas': 'kansas',
    'kentucky': 'kentucky',
    'louisiana': 'louisiana',
    'maine': 'maine',
    'maryland': 'maryland',
    'massachusetts': 'massachusetts',
    'michigan': 'michigan',
    'minnesota': 'minnesota',
    'mississippi': 'mississippi',
    'missouri': 'missouri',
    'montana': 'montana',
    'nebraska': 'nebraska',
    'nevada': 'nevada',
    'new hampshire': 'new-hampshire',
    'new jersey': 'new-jersey',
    'new mexico': 'new-mexico',
    'new york': 'new-york',
    'north carolina': 'north-carolina',
    'north dakota': 'north-dakota',
    'ohio': 'ohio',
    'oklahoma': 'oklahoma',
    'oregon': 'oregon',
    'pennsylvania': 'pennsylvania',
    'rhode island': 'rhode-island',
    'south carolina': 'south-carolina',
    'south dakota': 'south-dakota',
    'tennessee': 'tennessee',
    'texas': 'texas',
    'utah': 'utah',
    'vermont': 'vermont',
    'virginia': 'virginia',
    'washington': 'washington',
    'west virginia': 'west-virginia',
    'wisconsin': 'wisconsin',
    'wyoming': 'wyoming',
}

# States to process (50 states)
ALL_STATES = list(STATE_SLUGS.keys())


def extract_chambers_from_html(
    html: str,
    state_name: str,
    allow_website_guess: bool = False,
    show_progress: bool = False,
) -> List[Dict]:
    """Extracts chamber of commerce information from HTML.

    The page structure contains:
    - h3 tags with chamber names
    - Followed by p tag(s) with location info

    Args:
        html (str): The HTML content to parse
        state_name (str): The state name for reference
        allow_website_guess (bool): If True, attempts to guess/lookup missing websites.
        show_progress (bool): If True, shows a progress bar while processing chambers.

    Returns:
        List[Dict]: List of chamber dictionaries with keys:
                   - chamber_name: Name of the chamber
                   - city: City location
                   - state: State abbreviation
                   - website: Primary non-maps link for the chamber
                   - website_guess: Website found via guessing/search (if any)
                   - website_source: "scraped" | "guessed" | "none"
                   - maps_link: Google Maps / map link when present
    """
    soup = BeautifulSoup(html, 'html.parser')
    chambers = []
    seen_keys = set()
    
    # State abbreviation mapping for all US states
    state_abbr_map = {
        'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
        'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
        'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
        'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
        'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
        'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
        'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
        'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
        'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
        'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
        'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
        'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
        'wisconsin': 'WI', 'wyoming': 'WY',
    }
    
    state_name_lower = state_name.lower()
    expected_state_abbr = state_abbr_map.get(state_name_lower, state_name.upper()[:2])
    
    # Find all h3 tags which contain chamber names
    h3_tags = soup.find_all('h3')

    iterable = h3_tags
    if show_progress and TQDM_AVAILABLE:
        iterable = tqdm(h3_tags, desc=f"{state_name.title()} chambers", unit="chamber")

    for h3 in iterable:
        chamber_name = h3.get_text(strip=True)
        
        # Skip empty or too short names
        if not chamber_name or len(chamber_name) < 2:
            continue
        
        # Skip navigation/header elements
        skip_keywords = ['search', 'filter', 'by city', 'latest on', 'additional links', 'all states', 'welcome to']
        if any(keyword in chamber_name.lower() for keyword in skip_keywords):
            continue
        
        # Get the deduplication key to avoid duplicates
        key = dedupe_key(chamber_name)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        
        # Extract location info from following p tags
        city = ""

        # The structure alternates between:
        # - h3: Chamber Name
        # - p: Either membership status OR city info
        # - p: (optional) the other piece of info

        current = h3.next_sibling
        p_texts = []

        # Collect up to 3 p tags following the h3
        while current and len(p_texts) < 3:
            if isinstance(current, str):
                if current.strip():
                    pass  # Skip standalone text
            elif hasattr(current, 'name') and current.name == 'p':
                text = current.get_text(strip=True)
                if text:
                    p_texts.append(text)
            elif hasattr(current, 'name') and current.name == 'h3':
                # Hit the next chamber, stop
                break

            current = current.next_sibling

        # Parse collected text
        for text in p_texts:
            if ',' in text and len(text) < 50:  # Location format: "City, ST"
                city = text

        # Look inside the nearest parent container for more data (links, phone, email, address)
        website = ''
        website_guess = ''
        website_source = ''
        maps_link = ''
        all_links = []
        phone = ''
        email = ''
        address = ''
        image = ''

        # Prefer the enclosing <details> block which contains expanded card info
        parent = h3.find_parent('details')
        if not parent:
            parent = h3.find_parent()
        if parent:
            # collect links
            for a in parent.find_all('a', href=True):
                href = a['href'].strip()
                if href and href not in all_links:
                    all_links.append(href)

            # Deduce which links are map links vs website links.
            # If all_links contains a maps link + a website, we want each in its own column.
            for href in all_links:
                if not href.startswith('http'):
                    continue

                if any(skip in href for skip in ['maps.google.', 'google.com/maps', 'bing.com/maps', 'openstreetmap.org']):
                    # Prefer first maps link
                    if not maps_link:
                        maps_link = normalize_site(href)
                    continue

                # Skip internal uschamber pages
                if 'uschamber.com' in href:
                    continue

                # First non-maps external link becomes the website
                if not website:
                    website = normalize_site(href)
                website_source = 'scraped'

        # extract email and phone from parent text
        parent_text = parent.get_text(' ', strip=True) if parent else ''
        # email
        m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", parent_text)
        if m:
            email = m.group(0)
        # phone
        p = re.search(r"\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", parent_text)
        if p:
            phone = p.group(0)
        # address: look for street patterns
        addr_match = re.search(r"\d+\s+[\w\s.#,-]{5,100}", parent_text)
        if addr_match:
            address = addr_match.group(0).strip()
            # Strip trailing "Website ..." if the address text included embedded website metadata
            if "Website" in address:
                address = address.split("Website", 1)[0].strip()
        # image
        img = parent.find('img') if parent else None
        if img and img.get('src'):
            image = img.get('src')

            # extract email and phone from parent text
            parent_text = parent.get_text(' ', strip=True)
            # email
            m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", parent_text)
            if m:
                email = m.group(0)
            # phone
            p = re.search(r"\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", parent_text)
            if p:
                phone = p.group(0)
            # address: look for street patterns
            addr_match = re.search(r"\d+\s+[\w\s.#,-]{5,100}", parent_text)
            if addr_match:
                address = addr_match.group(0).strip()
                # Strip trailing "Website ..." if the address text included embedded website metadata
                if "Website" in address:
                    address = address.split("Website", 1)[0].strip()
            # image
            img = parent.find('img')
            if img and img.get('src'):
                image = img.get('src')
        
        # Extract state abbreviation from city if present
        state_abbr = expected_state_abbr
        if city:
            # Extract state code from "City, ST" format
            match = re.search(r',\s*([A-Z]{2})\s*$', city)
            if match:
                state_abbr = match.group(1)
        
        # If we didn't find a website link, optionally try a few common guesses and validate them
        if not website and allow_website_guess:
            website_guess = guess_website_from_name(
                chamber_name,
                state_abbr=state_abbr,
                address=address,
                max_attempts=5,
                timeout=4,
                allow_search=True,
            )
            if website_guess:
                website = website_guess
                website_source = 'guessed'

        # Mark entries that still have no website
        if not website:
            website_source = 'none'
        elif not website_source:
            # if website arrived but source wasn't set, assume scraped
            website_source = 'scraped'

        # Only add if we have a chamber name (city is optional in case of data issues)
        if chamber_name:
            chambers.append({
                'chamber_name': chamber_name,
                'city': city,
                'state': state_abbr,
                'website': website,
                'website_guess': website_guess,
                'website_source': website_source,
                'maps_link': maps_link,
                'all_links': ';'.join(all_links),
                'phone': phone,
                'email': email,
                'address': address,
                'image': image
            })
    
    return chambers


# Simple rate limiting to avoid hitting DuckDuckGo too fast.
_last_ddg_search_time = 0.0
_DDG_MIN_INTERVAL = 2.0  # seconds


def _extract_duckduckgo_result_urls(query: str, max_results: int = 3) -> list[str]:
    """Search DuckDuckGo (HTML endpoint) and return a list of result URLs.

    Uses the DuckDuckGo HTML search page and extracts the redirected URLs from the
    `uddg` parameter in the result links.
    """
    global _last_ddg_search_time

    # Rate limit / avoid being blocked.
    # DuckDuckGo may throttle heavy scraping; keep at least 2 seconds between requests.
    elapsed = time.time() - _last_ddg_search_time
    if elapsed < _DDG_MIN_INTERVAL:
        time.sleep(_DDG_MIN_INTERVAL - elapsed)

    try:
        from urllib.parse import quote
        encoded_query = quote(query)
        html = fetch_with_retries(
            f"https://duckduckgo.com/html/?q={encoded_query}",
            timeout=15,
            retries=2,
        )
        _last_ddg_search_time = time.time()
    except Exception as e:
        return []

    soup = BeautifulSoup(html, 'html.parser')
    urls = []
    for a in soup.select('a.result__a'):
        href = a.get('href', '')
        if not href:
            continue
        if href.startswith('//'):
            href = 'https:' + href

        # DuckDuckGo wraps results in a redirect link with uddg parameter.
        # Example: https://duckduckgo.com/l/?uddg=https%3A%2F%2Fwww.anchoragechamber.org%2F&...
        parsed = urlparse(href)
        qs = parse_qs(parsed.query)
        if 'uddg' in qs and qs['uddg']:
            target = unquote(qs['uddg'][0])
        else:
            # Fallback: if href looks like a real URL
            target = href if href.startswith('http') else ''

        if target and target not in urls:
            urls.append(target)
            if len(urls) >= max_results:
                break

    return urls


def _generate_domain_candidates(chamber_name: str) -> list:
    """Generate prioritized domain candidates from chamber name.
    
    Strategy:
      1) Remove common stopwords (county, city, parish, chamber, commerce, etc)
      2) Build candidates in priority order:
         - Keywords with 'chamber' suffix (.com, .org)
         - Full slugs with 'chamber' (.com, .org)
         - Keywords (.com, .org)
         - Full slugs (.com, .org)
    """
    if not chamber_name:
        return []
    
    stopwords = {'county', 'city', 'parish', 'chamber', 'of', 'commerce', 'and', 'the', 'a', 'an'}
    
    # Create normalized versions
    name_lower = chamber_name.lower().strip()
    alnum_slug = re.sub(r'[^a-z0-9]+', '', name_lower)
    dash_slug = re.sub(r'[^a-z0-9]+', '-', name_lower).strip('-')
    
    # Extract keywords (words between spaces/hyphens)
    keywords = re.split(r'[\s\-]+', name_lower)
    keywords = [k for k in keywords if k and k not in stopwords and len(k) > 2]
    
    candidates = []
    seen = set()
    
    # Priority 1: Keywords with 'chamber' suffix
    for kw in keywords:
        for suffix in ['chamber', 'chamberofcommerce']:
            for tld in ['.com', '.org']:
                for prefix in ['', 'www.']:
                    url = f"https://{prefix}{kw}{suffix}{tld}"
                    if url not in seen:
                        candidates.append(url)
                        seen.add(url)
    
    # Priority 2: Keywords with hyphen + 'chamber'
    for kw in keywords:
        for suffix in ['chamber', 'chamber-of-commerce']:
            for tld in ['.com', '.org']:
                for prefix in ['', 'www.']:
                    url = f"https://{prefix}{kw}-{suffix}{tld}"
                    if url not in seen:
                        candidates.append(url)
                        seen.add(url)
    
    # Priority 3: Full slugs with chamber suffix
    for suffix in ['chamber', 'chamberofcommerce']:
        for tld in ['.com', '.org']:
            for prefix in ['', 'www.']:
                for slug in [alnum_slug, dash_slug]:
                    if slug:
                        url = f"https://{prefix}{slug}{suffix}{tld}"
                        if url not in seen:
                            candidates.append(url)
                            seen.add(url)
    
    # Priority 4: Full slugs direct
    for tld in ['.com', '.org']:
        for prefix in ['', 'www.']:
            for slug in [alnum_slug, dash_slug]:
                if slug:
                    url = f"https://{prefix}{slug}{tld}"
                    if url not in seen:
                        candidates.append(url)
                        seen.add(url)
    
    # Priority 5: Keywords direct
    for tld in ['.com', '.org']:
        for prefix in ['', 'www.']:
            for kw in keywords:
                url = f"https://{prefix}{kw}{tld}"
                if url not in seen:
                    candidates.append(url)
                    seen.add(url)
    
    return candidates


def _validate_with_address(url: str, address: str = '', city: str = '', state_abbr: str = '', html_content: str = '') -> bool:
    """Validate that a URL is the right chamber by checking for address components.
    
    Uses the provided HTML content if available (avoids double-fetching).
    STRICT: Requires at least address components (city OR zip) to match.
    This prevents false positives from redirects or unrelated sites.
    """
    if not url:
        return False
    
    # If HTML not provided, try to fetch it (with very short timeout)
    if not html_content:
        try:
            html_content = fetch_with_retries(url, timeout=1, retries=0)
        except Exception:
            return False
    
    if not html_content or len(html_content) < 100:
        return False  # Page too small or unavailable
    
    # Don't process huge pages (likely redirects or generic sites like ashford.com)
    if len(html_content) > 1500000:  # 1.5MB limit
        return False
    
    text = html_content.lower()
    
    # Extract validation components - be strict
    location_checks = []
    
    # Primary: City from address
    if city:
        location_checks.append(city.lower())
    elif address and ',' in address:
        city_part = address.split(',', 1)[0].strip()
        if city_part and len(city_part) > 2:
            location_checks.append(city_part.lower())
    
    # Secondary: Zip code
    zip_code = None
    if address:
        zip_match = re.search(r'\b\d{5}\b', address)
        if zip_match:
            zip_code = zip_match.group(0)
            location_checks.append(zip_code)
    
    # STRICT validation: Must have at least city OR zip to check
    if not location_checks:
        return False  # No identifying info - reject to be safe
    
    # Check if ANY location identifier appears in the page
    found_match = any(check in text for check in location_checks)
    
    # If no match, definitely reject (site is not about this location)
    if not found_match:
        return False
    
    return True


def _is_site_relevant(url: str, state_abbr: str, address: str, chamber_name: str = '') -> bool:
    """Checks if a fetched site appears to match the expected state/address.

    First checks the domain/URL for relevance, then validates content if needed.
    Accepts directory sites like chamberofcommerce.com automatically.
    """
    if not url:
        return False

    url_lower = url.lower()
    
    # Accept known directory sites that are generally legitimate
    if 'chamberofcommerce.com' in url_lower:
        return True
    if 'businessdirectory' in url_lower:
        return True
    
    # Check if domain contains "chamber" keyword - strong signal
    if 'chamber' in url_lower:
        return True

    try:
        html = fetch_with_retries(url, timeout=7, retries=1)
    except Exception:
        return False

    text = html.lower()
    # Check for state abbreviation, city or zip, if available.
    state_abbr = (state_abbr or '').strip().lower()
    address = (address or '').strip().lower()

    checks = []
    if state_abbr:
        checks.append(state_abbr)

    # Try to use the city portion of the address (before comma)
    if ',' in address:
        city = address.split(',', 1)[0].strip()
        if city:
            checks.append(city)

    # Also try to include zip code information if present
    zip_match = re.search(r"\b\d{5}\b", address)
    if zip_match:
        checks.append(zip_match.group(0))

    # If no metadata, conservatively accept (page loaded successfully)
    if not checks:
        return True

    # Require at least one match to consider the site relevant
    return any(check in text for check in checks)


def guess_website_from_name(
    chamber_name: str,
    state_abbr: str = '',
    address: str = '',
    max_attempts: int = 3,
    timeout: int = 3,
    allow_search: bool = True,
) -> str:
    """Try to find a working website URL for a chamber.

    Strategy:
      1) Generate smart domain candidates based on chamber name and test them
      2) Validate by checking for address components (city, zip, etc) on the page
      3) Skip timeouts and move to next candidate
    """
    if not chamber_name:
        return ''

    # Generate prioritized domain candidates
    candidates = _generate_domain_candidates(chamber_name)
    
    # Extract city from address for validation
    city = ''
    if address and ',' in address:
        city = address.split(',', 1)[0].strip()
    
    # Test candidates - check domain for chamber/commerce keywords
    for url in candidates[:max_attempts]:
        try:
            # Fetch with timeout - skip if too slow or too large
            html = fetch_with_retries(url, timeout=timeout, retries=1)
            
            # Skip huge pages (likely not a chamber site)
            if html and len(html) > 1500000:  # 1.5MB limit
                continue
            
            url_lower = url.lower()
            
            # Strong domain signals - must have chamber/commerce keywords
            if 'chamber' in url_lower or 'commerce' in url_lower or 'business' in url_lower or 'directory' in url_lower:
                # Strict validation: check address components actually match
                # Use a short timeout on validation to prevent hanging
                try:
                    if _validate_with_address(url, address, city, state_abbr, html):
                        return url
                except Exception:
                    # Validation failed or timed out - skip
                    continue
        except Exception:
            # Timeout or error on fetch - skip this candidate and try next one
            continue

    # If we still don't have a valid site and searching is allowed, try DuckDuckGo results
    if allow_search:
        # Try multiple query variations, starting with simpler ones
        query_variations = []
        # Simple: just chamber name + state abbr
        if state_abbr:
            query_variations.append(f"{chamber_name} {state_abbr}")
        query_variations.append(f"{chamber_name} chamber")
        query_variations.append(chamber_name)
        
        # Extract city from address if available
        if address and ',' in address:
            city_part = address.split(',', 1)[0].strip()
            if city_part:
                query_variations.append(f"{city_part} chamber {state_abbr or ''}")
                query_variations.append(f"{city_part} chamber of commerce")
        
        for query in query_variations:
            urls = _extract_duckduckgo_result_urls(query, max_results=max_attempts)
            if urls:
                for url in urls:
                    # Slow down queries to avoid rate limiting
                    time.sleep(0.5)
                    if _validate_with_address(url, address, city, state_abbr):
                        return url

    return ''


def find_member_directory_url(
    seed_url: str,
    max_pages: int = 5,
    timeout: int = 15,
    verbose: bool = False,
) -> str:
    """Locate a member/business directory URL starting from a seed site.

    Args:
        seed_url: Starting URL (usually the chamber's website homepage)
        max_pages: Max number of pages to visit during discovery
        timeout: Request timeout seconds
        verbose: If True, print detailed heuristic decisions.

    Returns:
        The first discovered member directory URL or empty string if none found.
    """
    if not seed_url:
        return ''

    # Ensure we have a normalized absolute URL
    seed_url = normalize_site(seed_url)
    parsed = urlparse(seed_url)
    if not parsed.scheme or not parsed.netloc:
        return ''

    base_host = parsed.netloc

    # Keywords to detect candidate links
    keywords = [
        'member', 'members', 'member-directory', 'memberdirectory',
        'directory', 'business-directory', 'businessdirectory',
        'find-members', 'findmembers', 'near-me', 'search'
    ]

    def is_candidate_link(href: str, text: str) -> bool:
        if not href:
            return False
        href_lower = href.lower()
        text_lower = (text or '').lower()
        # Prefer links with keywords in URL or link text
        return any(k in href_lower for k in keywords) or any(k in text_lower for k in keywords)

    def normalize_href(href: str) -> str:
        if not href:
            return ''
        href = href.strip()
        if href.startswith('javascript:') or href.startswith('#'):
            return ''
        return normalize_site(href, base=seed_url)

    def _page_has_az_index(soup: BeautifulSoup) -> bool:
        """Detect A-Z / 0-9 index navigation links."""
        seen = set()
        for a in soup.find_all('a'):
            txt = a.get_text(' ', strip=True)
            if not txt:
                continue
            txt = txt.strip()
            if re.fullmatch(r'[A-Za-z]', txt):
                seen.add(txt.upper())
            elif re.fullmatch(r'0\s*[-–—]\s*9', txt) or txt.lower() in ('0-9', '0–9', '0—9'):
                seen.add('0-9')
        return len(seen) >= 6

    def _page_has_category_listing(soup: BeautifulSoup) -> bool:
        """Detect category-style directory listings (header + list of links)."""
        groups = 0
        for heading in soup.find_all(['h1', 'h2', 'h3', 'h4']):
            next_el = heading.find_next_sibling()
            if not next_el:
                continue

            if next_el.name in ('ul', 'ol'):
                links = next_el.find_all('a')
                if len(links) >= 2:
                    groups += 1
            elif next_el.name == 'div':
                links = next_el.find_all('a')
                if len(links) >= 3:
                    groups += 1

            if groups >= 3:
                return True

        return False

    def _get_directory_hints(html: str, url: str) -> List[str]:
        soup = BeautifulSoup(html, 'html.parser')
        hints: List[str] = []

        url_lower = (url or '').lower()
        if 'directory' in url_lower:
            hints.append("url contains 'directory'")
        if 'member' in url_lower:
            hints.append("url contains 'member'")

        if _page_has_az_index(soup):
            hints.append('A–Z index')
        if _page_has_category_listing(soup):
            hints.append('category listing')

        return hints

    visited = set()
    to_visit = [seed_url]

    def scan_page(html: str, current_url: str) -> Optional[str]:
        hints = _get_directory_hints(html, current_url)
        if hints:
            if verbose:
                print(f"  → {current_url} looks like a directory ({'; '.join(hints)})")
            return current_url

        soup = BeautifulSoup(html, 'html.parser')
        candidates = []
        for a in soup.find_all('a', href=True):
            href = normalize_href(a['href'])
            if not href:
                continue
            if href in visited:
                continue

            text = a.get_text(' ', strip=True)
            if is_candidate_link(href, text):
                # Prefer links that are clearly directory-like
                candidates.append(href)

        # Try to return the best candidate
        if candidates:
            # Prefer candidates that contain 'member' or 'directory' in path
            for c in candidates:
                if any(k in c.lower() for k in ['member', 'directory']):
                    if verbose:
                        print(f"  → Choosing candidate {c} (keyword match)")
                    return c
            if verbose:
                print(f"  → Choosing candidate {candidates[0]} (fallback)")
            return candidates[0]

        return None

    pages_visited = 0
    while to_visit and pages_visited < max_pages:
        url = to_visit.pop(0)
        if url in visited:
            continue
        visited.add(url)

        try:
            html = fetch_with_retries(url, timeout=timeout, retries=2)
        except Exception:
            continue

        pages_visited += 1

        found = scan_page(html, url)
        if found:
            return found

        # Queue additional internal links for exploration
        soup = BeautifulSoup(html, 'html.parser')
        for a in soup.find_all('a', href=True):
            href = normalize_href(a['href'])
            if not href or href in visited:
                continue
            parsed_href = urlparse(href)
            if parsed_href.netloc != base_host:
                continue
            # Ignore mailto and similar
            if parsed_href.scheme not in ('http', 'https'):
                continue
            to_visit.append(href)

    return ''


def get_state_chambers(
    state_name: str,
    snapshot_path: Optional[str] = None,
    allow_website_guess: bool = False,
    show_progress: bool = False,
) -> List[Dict]:
    """
    Extracts chamber of commerce information for a given state from uschamber.com
    
    Args:
        state_name (str): The name of the state to process
        snapshot_path (str, optional): Path to a local HTML snapshot
        
    Returns:
        list: A list of dictionaries, each representing a chamber of commerce
    """
    state_lower = state_name.lower()
    
    if snapshot_path:
        # Load from local snapshot
        with open(snapshot_path, 'r', encoding='utf-8') as f:
            html = f.read()
        print(f"✓ Loaded snapshot for {state_name} from {snapshot_path}")
    else:
        # Fetch from web
        state_slug = STATE_SLUGS.get(state_lower, state_lower)
        state_url = f"{USCHAMBER_BASE_URL}/{state_slug}"
        print(f"→ Fetching {state_name} from {state_url}...")
        
        try:
            html = fetch_with_retries(state_url, timeout=15, retries=3)
            if not html:
                print(f"✗ Failed to fetch data for {state_name}")
                return []
            print(f"✓ Retrieved {len(html)} bytes for {state_name}")
        except Exception as e:
            print(f"✗ Error fetching {state_name}: {e}")
            return []
    
    # Extract chambers from HTML
    chambers = extract_chambers_from_html(
        html,
        state_name,
        allow_website_guess=allow_website_guess,
        show_progress=show_progress,
    )
    print(f"✓ Extracted {len(chambers)} chambers from {state_name}")
    
    return chambers


def _derive_seed_url_from_row(row: Dict) -> str:
    """Pick a starting URL for directory discovery."""
    # Prefer the chamber's website (non-Google-maps link)
    seed = (row.get('website') or '').strip()
    if seed:
        return seed

    # Next, use the explicit maps link (if available) as a last resort
    maps_link = (row.get('maps_link') or '').strip()
    if maps_link:
        return maps_link

    # Fallback: use first absolute link from all_links
    for link in (row.get('all_links') or '').split(';'):
        link = link.strip()
        if link.startswith('http'):
            return link
    return ''


def enrich_with_member_directory_links(
    chambers: List[Dict],
    max_pages: int = 5,
    verbose: bool = True,
    progress_interval: int = 10,
) -> List[Dict]:
    """Add a member_directory_url field to each chamber row.

    Args:
        chambers: List of chamber records
        max_pages: Max pages to crawl per domain
        verbose: Whether to print progress updates
        progress_interval: How many chambers between progress prints
    """
    # Cache by host to avoid repeated crawling for the same domain.
    host_cache: Dict[str, str] = {}

    total = len(chambers)
    for idx, chamber in enumerate(chambers, start=1):
        chamber.setdefault('member_directory_url', '')
        if chamber.get('member_directory_url'):
            continue

        seed_url = _derive_seed_url_from_row(chamber)
        if not seed_url:
            continue

        parsed = urlparse(seed_url)
        host = parsed.netloc
        if not host:
            continue

        if host in host_cache:
            chamber['member_directory_url'] = host_cache[host]
            continue

        member_dir_url = find_member_directory_url(
            seed_url, max_pages=max_pages, verbose=verbose
        )
        chamber['member_directory_url'] = member_dir_url or ''
        host_cache[host] = chamber['member_directory_url']

        if verbose and idx % progress_interval == 0:
            print(f"    → {idx}/{total} chambers processed (last host: {host})")

    return chambers


def _append_df_to_xlsx(df: pd.DataFrame, xlsx_path: str, sheet_name: str) -> None:
    """Append a DataFrame into an XLSX sheet (create/append)."""
    from openpyxl import load_workbook, Workbook

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # If sheet has no header/values, write our header
    def _sheet_has_header(_ws):
        # Consider a sheet "empty" if the first row has no non-empty cells
        for row in _ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if any(cell not in (None, '') for cell in row):
                return True
        return False

    if not _sheet_has_header(ws):
        ws.append(list(df.columns))

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    # Remove default sheet if still empty and not needed
    if 'Sheet' in wb.sheetnames:
        ws_def = wb['Sheet']
        # If sheet has no values (all cells empty), remove it.
        if all(cell.value is None for row in ws_def.iter_rows(max_row=1) for cell in row):
            del wb['Sheet']

    # Make sure the sheet we wrote is the active one
    if sheet_name in wb.sheetnames:
        wb.active = wb.sheetnames.index(sheet_name)

    wb.save(xlsx_path)


def _write_df_to_xlsx_sheet(df: pd.DataFrame, xlsx_path: str, sheet_name: str) -> None:
    """Overwrite a sheet in an XLSX (create new workbook if needed)."""
    from openpyxl import load_workbook, Workbook

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    # Remove default sheet if still unused
    if 'Sheet' in wb.sheetnames:
        ws_def = wb['Sheet']
        if all(cell.value is None for row in ws_def.iter_rows(max_row=1) for cell in row):
            del wb['Sheet']

    # Make sure the sheet we wrote is the active one
    if sheet_name in wb.sheetnames:
        wb.active = wb.sheetnames.index(sheet_name)

    wb.save(xlsx_path)


def scrape_all_states(
    output_file: str = "uschamber_all_states.csv",
    single_sheet: bool = True,
    discover_member_links: bool = False,
    member_link_max_pages: int = 5,
    allow_website_guess: bool = False,
) -> None:
    """Scrapes all states and writes progress incrementally to CSV and XLSX."""
    failed_states = []
    successful_states = []

    output_path = os.path.join('output', output_file)
    xlsx_path = os.path.join('output', os.path.splitext(output_file)[0] + '.xlsx')

    # Ensure output directory exists
    os.makedirs('output', exist_ok=True)

    # Reset output files so we can append reliably
    if os.path.exists(output_path):
        os.remove(output_path)
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    print(f"\n{'='*70}")
    print(f"  US CHAMBER OF COMMERCE SCRAPER - ALL STATES")
    print(f"  Target: {len(ALL_STATES)} states")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    total_chambers = 0
    for i, state in enumerate(ALL_STATES, 1):
        prefix = f"[{i}/{len(ALL_STATES)}]"
        print(f"\n{prefix} Processing {state.title()}...")
        try:
            chambers = get_state_chambers(
                state,
                allow_website_guess=allow_website_guess,
                show_progress=True,
            )
            if not chambers:
                failed_states.append(state)
                continue

            if discover_member_links:
                print("  → Discovering member directory links (this may take a while)...")
                chambers = enrich_with_member_directory_links(
                    chambers, max_pages=member_link_max_pages
                )

            df_state = pd.DataFrame(chambers)
            df_state = df_state.sort_values(['state', 'chamber_name']).reset_index(drop=True)
            df_state = df_state.reindex(columns=OUTPUT_COLUMNS).fillna('')

            # Append to CSV (write header if first state)
            write_header = not os.path.exists(output_path)
            df_state.to_csv(output_path, mode='a', index=False, header=write_header)

            # Append/Write to XLSX
            if single_sheet:
                _append_df_to_xlsx(df_state, xlsx_path, sheet_name='AllStates')
            else:
                sheet_name = state.title()[:31].replace('/', '_').replace('\\', '_')
                _write_df_to_xlsx_sheet(df_state, xlsx_path, sheet_name=sheet_name)

            successful_states.append(state)
            total_chambers += len(chambers)
            print(f"  → State complete: {len(chambers)} chambers (total so far: {total_chambers})")

        except Exception as e:
            print(f"    ✗ Exception: {e}")
            failed_states.append(state)

    print(f"\n{'='*70}")
    print(f"  SCRAPING COMPLETE")
    print(f"  Total chambers found: {total_chambers}")
    print(f"  Successful states: {len(successful_states)}")
    print(f"  Failed states: {len(failed_states)}")
    print(f"  Output file: {output_path}")
    print(f"  Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    if failed_states:
        print(f"Failed states: {', '.join(failed_states)}\n")


def fill_missing_websites_in_output(
    input_csv: str,
    output_csv: str,
    allow_search: bool = True,
    max_attempts: int = 5,
    timeout: int = 4,
) -> None:
    """Fill missing website values in an existing output CSV/XLSX.

    This reads the existing CSV, attempts to guess/lookup only those rows
    where "website" is empty, and writes updates INCREMENTALLY to CSV + XLSX
    as each row is processed.
    """
    if not os.path.exists(input_csv):
        print(f"Error: input file not found: {input_csv}", flush=True)
        return

    print(f"Loading existing output from {input_csv}...", flush=True)
    df = pd.read_csv(input_csv)

    missing_mask = df['website'].isna() | (df['website'].astype(str).str.strip() == '')
    missing_count = int(missing_mask.sum())
    print(f"Found {missing_count} rows with empty website (out of {len(df)}).", flush=True)

    if missing_count == 0:
        print("No missing websites found; nothing to do.", flush=True)
        return

    total_missing = missing_count
    progress = 0

    for idx, row in df[missing_mask].iterrows():
        chamber_name = row.get('chamber_name', '')
        state_abbr = row.get('state', '')
        address = row.get('address', '')
        city = row.get('city', '')  # Get city column directly

        print(f"Processing row {progress + 1}/{total_missing}: {chamber_name} ({state_abbr})", flush=True)

        guess = guess_website_from_name(
            chamber_name,
            state_abbr=state_abbr,
            address=address,
            city=city,
            max_attempts=max_attempts,
            timeout=timeout,
            allow_search=allow_search,
        )

        if guess:
            df.at[idx, 'website'] = guess
            df.at[idx, 'website_guess'] = guess
            df.at[idx, 'website_source'] = 'guessed'
            print(f"  ✓ Filled: {guess}", flush=True)
        else:
            df.at[idx, 'website_source'] = 'none'
            print(f"  ✗ No guess found", flush=True)

        progress += 1
        
        # Write incrementally every row (important: don't wait until end)
        try:
            df_to_write = df.reindex(columns=OUTPUT_COLUMNS).fillna('')
            df_to_write.to_csv(input_csv, index=False)
            
            # Also update XLSX
            try:
                from openpyxl import load_workbook, Workbook
                if os.path.exists(output_csv):
                    wb = load_workbook(output_csv)
                else:
                    wb = Workbook()
                
                sheet_name = 'AllStates'
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(sheet_name)
                ws.append(list(df_to_write.columns))
                for row_data in df_to_write.itertuples(index=False, name=None):
                    ws.append(list(row_data))
                
                if 'Sheet' in wb.sheetnames:
                    ws_default = wb['Sheet']
                    if all(cell.value is None for row in ws_default.iter_rows(max_row=1) for cell in row):
                        del wb['Sheet']
                
                if sheet_name in wb.sheetnames:
                    wb.active = wb.sheetnames.index(sheet_name)
                wb.save(output_csv)
            except Exception:
                pass  # If XLSX fails, at least CSV is saved
        except Exception as e:
            print(f"  Warning: Failed to write update: {e}", flush=True)
        
        if progress % 25 == 0 or progress == total_missing:
            print(f"  Progress: {progress}/{total_missing} missing websites...", flush=True)

    print(f"All {total_missing} rows processed. Output saved.", flush=True)


def main():
    parser = argparse.ArgumentParser(
        description='Scrape US Chamber of Commerce directories'
    )
    parser.add_argument(
        '--state',
        type=str,
        help='Scrape specific state (e.g., alaska, california)'
    )
    parser.add_argument(
        '--snapshot',
        type=str,
        help='Path to local HTML snapshot file'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='uschamber_all_states.csv',
        help='Output CSV file name (in output/ directory)'
    )
    parser.add_argument(
        '--single-sheet',
        dest='single_sheet',
        action='store_true',
        default=True,
        help='Write all states into a single XLSX sheet (default)'
    )
    parser.add_argument(
        '--multi-sheet',
        dest='single_sheet',
        action='store_false',
        help='Write one XLSX sheet per state (optional)'
    )
    parser.add_argument(
        '--discover-members',
        dest='discover_members',
        action='store_true',
        default=False,
        help='Attempt to discover member/directory URLs for each chamber (may take significant time)'
    )
    parser.add_argument(
        '--member-link-pages',
        dest='member_link_pages',
        type=int,
        default=5,
        help='Max pages to crawl per site when discovering member directory links'
    )
    parser.add_argument(
        '--augment',
        dest='augment',
        action='store_true',
        default=False,
        help='Augment an existing output CSV/XLSX by discovering member directory URLs (uses existing output file)'
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Scrape all states (default)'
    )
    parser.add_argument(
        '--guess-websites',
        dest='guess_websites',
        action='store_true',
        default=False,
        help='Attempt to guess/lookup missing websites (slower but may fill in blanks)'
    )
    parser.add_argument(
        '--fill-missing-websites',
        dest='fill_missing_websites',
        action='store_true',
        default=False,
        help='Load existing output and only attempt to fill blank website fields (does not re-scrape all states)'
    )
    
    args = parser.parse_args()

    if args.fill_missing_websites:
        input_csv = os.path.join('output', args.output)
        output_xlsx = os.path.join('output', os.path.splitext(args.output)[0] + '.xlsx')
        fill_missing_websites_in_output(
            input_csv,
            output_xlsx,
            allow_search=False,  # Disable DuckDuckGo search due to timeouts, rely on domain guessing
            max_attempts=20,  # Try more candidates before giving up
        )
        return

    if args.augment:
        # Use existing output file and add member directory links
        input_csv = os.path.join('output', args.output)
        if not os.path.exists(input_csv):
            print(f"Error: Output file not found at {input_csv}. Run scrape first.")
            sys.exit(1)

        print(f"Loading existing output from {input_csv}...")
        df = pd.read_csv(input_csv)
        chambers = df.to_dict(orient='records')

        print("\nDiscovering member directory links for existing records...")
        chambers = enrich_with_member_directory_links(
            chambers, max_pages=args.member_link_pages
        )

        df = pd.DataFrame(chambers)
        df = df.sort_values(['state', 'chamber_name']).reset_index(drop=True)
        df = df.reindex(columns=OUTPUT_COLUMNS).fillna('')

        # Save updated CSV + XLSX
        df.to_csv(input_csv, index=False)
        xlsx_path = os.path.join('output', os.path.splitext(args.output)[0] + '.xlsx')
        try:
            with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                sheet_name = 'AllStates'
                sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception:
            try:
                df.to_excel(xlsx_path, index=False)
            except Exception:
                pass

        print(f"\nAugmented data saved to {input_csv} and {xlsx_path}")
        return

    if args.state:
        state_lower = args.state.lower()
        if state_lower not in STATE_SLUGS:
            print(f"Error: Unknown state '{args.state}'")
            print(f"Available states: {', '.join(sorted(STATE_SLUGS.keys()))}")
            sys.exit(1)
        
        chambers = get_state_chambers(
            state_lower,
            snapshot_path=args.snapshot,
            allow_website_guess=args.guess_websites,
            show_progress=True,
        )
        
        if chambers:
            if args.discover_members:
                print("\nDiscovering member directory links for extracted chambers...")
                chambers = enrich_with_member_directory_links(
                    chambers, max_pages=args.member_link_pages
                )

            df = pd.DataFrame(chambers)
            df = df.reindex(columns=OUTPUT_COLUMNS).fillna('')
            print("\nExtracted chambers:")
            print(df.to_string())

            # Save to output (CSV + XLSX single-sheet)
            os.makedirs('output', exist_ok=True)
            csv_path = os.path.join('output', f'uschamber_{state_lower}.csv')
            xlsx_path = os.path.join('output', f'uschamber_{state_lower}.xlsx')
            df.to_csv(csv_path, index=False)
            try:
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    sheet_name = state_lower.title()[:31]
                    writer.book  # ensure workbook creation
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                try:
                    df.to_excel(xlsx_path, index=False)
                except Exception:
                    pass
            print(f"\nSaved to {csv_path} and {xlsx_path}")
    else:
        # Default: scrape all states
        scrape_all_states(
            args.output,
            single_sheet=args.single_sheet,
            discover_member_links=args.discover_members,
            member_link_max_pages=args.member_link_pages,
            allow_website_guess=args.guess_websites,
        )


if __name__ == '__main__':
    main()
