#!/usr/bin/env python3
"""Scrape each state page for chamber names/links and label by town/county where possible.

Writes results to `output/state_chambers_detailed.xlsx` in sheet `chambers_detailed`.

Usage:
    python3 scripts/scrape_state_chambers_detailed.py --input output/state_chambers.xlsx
"""
import argparse
import logging
import re
import time
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing


def is_external_link(url: str) -> bool:
    if not url:
        return False
    url = url.strip()
    if url.startswith("mailto:") or url.startswith("javascript:"):
        return False
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return False
    if "officialusa.com" in parsed.netloc.lower():
        return False
    return True


def fetch(url: str) -> str:
    tries = 0
    backoff = 1.0
    while tries < 6:
        try:
            r = requests.get(url, timeout=20)
            r.raise_for_status()
            return r.text
        except requests.HTTPError as e:
            status = getattr(e.response, 'status_code', None)
            if status == 429:
                logging.warning("429 for %s, backing off %ss", url, backoff)
                time.sleep(backoff)
                backoff *= 2
                tries += 1
                continue
            raise
        except requests.RequestException:
            tries += 1
            logging.warning("Request failed for %s, retrying in %ss", url, backoff)
            time.sleep(backoff)
            backoff *= 2
    raise RuntimeError(f"Failed to fetch {url} after retries")


def extract_location_from_context(text: str) -> str:
    if not text:
        return ""
    # Prefer 'X County'
    m = re.search(r'([A-Z][\w .\'-]+County)', text)
    if m:
        return m.group(1)
    # Prefer 'City, State' like 'Springfield, Illinois' or 'Springfield, IL'
    m = re.search(r'([A-Z][\w .\'-]+),\s*([A-Z]{2}|[A-Z][a-z]+)', text)
    if m:
        return m.group(1)
    # fallback: take short prefix (first sentence-like chunk)
    txt = text.strip()
    if len(txt) > 120:
        # try split on separators
        for sep in [' - ', ' – ', '—', '|', '•', '\n']:
            if sep in txt:
                return txt.split(sep)[0].strip()
        return txt.split('.')[0].strip()[:80]
    return txt


def extract_chambers_from_state(html: str, base_url: str) -> list:
    soup = BeautifulSoup(html, 'html.parser')
    results = []
    seen = set()
    # Look for likely containers with lists: li, table rows, p, div
    anchors = soup.find_all('a', href=True)
    for a in anchors:
        href = a['href'].strip()
        full = urljoin(base_url, href)
        if not is_external_link(full):
            continue
        if full in seen:
            continue
        seen.add(full)
        name = a.get_text(' ', strip=True) or full
        # get surrounding context
        parent = a
        context_text = ''
        for _ in range(4):
            parent = parent.parent
            if parent is None:
                break
            if parent.name in ('li', 'p', 'td', 'tr', 'div'):
                context_text = parent.get_text(' ', strip=True)
                break
        if not context_text:
            # try previous sibling text
            prev = a.previous_sibling
            if prev and isinstance(prev, str):
                context_text = prev.strip()
        location = extract_location_from_context(context_text)
        results.append({
            'chamber_name': name,
            'chamber_url': full,
            'location_guess': location,
            'anchor_text': a.get_text(' ', strip=True),
            'context': context_text,
        })

    # Find plain-text chamber mentions (no anchor) in list items, table cells, paragraphs
    for tag in soup.find_all(['li', 'tr', 'td', 'p']):
        text = tag.get_text(' ', strip=True)
        if not text:
            continue
        # Heuristic: contains 'Chamber' or 'chamber' or 'Chamber of Commerce'
        if re.search(r'\b[Cc]hamber\b', text) and not tag.find('a'):
            # try to split into lines and extract likely name
            lines = [ln.strip() for ln in re.split(r"\n|\\r|\\u2022|-|–|—|\|", text) if ln.strip()]
            candidate = None
            for ln in lines:
                if re.search(r'\b[Cc]hamber\b', ln):
                    candidate = ln
                    break
            if not candidate:
                candidate = lines[0]
            # avoid duplicates by name
            key = (candidate.lower(), None)
            if any((candidate.lower() == (r.get('chamber_name') or '').lower() for r in results)):
                continue
            # Context and location guess
            context_text = tag.get_text(' ', strip=True)
            location = extract_location_from_context(context_text)
            results.append({
                'chamber_name': candidate,
                'chamber_url': '',
                'location_guess': location,
                'anchor_text': '',
                'context': context_text,
            })

    # For entries without a chamber_url, attempt to discover probable URL heuristically
    for entry in results:
        if entry.get('chamber_url'):
            continue
        name = entry.get('chamber_name') or ''
        discovered = find_link_for_name(name, soup, base_url)
        if discovered:
            entry['chamber_url'] = discovered
    return results


def _crawl_one_state(state: str, url: str) -> list:
    """Fetch a single state page and extract chambers. Returns list of dicts."""
    logging.info('Crawling %s -> %s', state, url)
    html = fetch(url)
    items = extract_chambers_from_state(html, url)
    # polite per-request delay to reduce rate-limiting
    time.sleep(1.0)
    return items


DEFAULT_OUTPUT = 'output/state_chambers_detailed.xlsx'
DEFAULT_WORKERS = 2
SLOW_DELAY = 1.0


def crawl_all_states(input_xlsx: str, output_xlsx: str = DEFAULT_OUTPUT):
    df_states = pd.read_excel(input_xlsx)
    urls = []
    for _, row in df_states.iterrows():
        state = row.get('state') or ''
        url = row.get('url')
        if not isinstance(url, str) or not url:
            continue
        urls.append((state, url))

    all_rows = []
    workers = DEFAULT_WORKERS
    logging.info('Crawling %d state pages with %d workers', len(urls), workers)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_crawl_one_state, state, url): (state, url) for state, url in urls}
        for fut in as_completed(futures):
            state, url = futures[fut]
            try:
                items = fut.result()
            except Exception as e:
                logging.warning('Failed crawling %s: %s', url, e)
                continue
            for it in items:
                it['state'] = state
                it['source'] = url
                all_rows.append(it)

    if not all_rows:
        logging.info('No chamber entries found')
        return

    df = pd.DataFrame(all_rows).drop_duplicates(subset=['chamber_url'])

    # write to output xlsx
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    try:
        wb = load_workbook(output_xlsx)
        if 'chambers_detailed' in wb.sheetnames:
            del wb['chambers_detailed']
        ws = wb.create_sheet('chambers_detailed')
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(output_xlsx)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'chambers_detailed'
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(output_xlsx)


def find_link_for_name(name: str, soup: BeautifulSoup, base_url: str) -> str:
    """Try to locate a URL for a plain-text chamber `name` by scanning anchors on the same page.

    This avoids external HEAD requests to reduce network load and rate-limiting. Returns the
    first matching external anchor URL that contains a token from the `name`, or empty string.
    """
    if not name:
        return ''
    tokens = [t.lower() for t in re.findall(r"[A-Za-z0-9]+", name) if len(t) > 2]
    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        full = urljoin(base_url, href)
        if not is_external_link(full):
            continue
        href_low = full.lower()
        text_low = (a.get_text(' ', strip=True) or '').lower()
        if any(tok in href_low or tok in text_low for tok in tokens[:3]):
            return full
    return ''


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', '-i', default='output/state_chambers.xlsx')
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    crawl_all_states(args.input)


if __name__ == '__main__':
    main()
