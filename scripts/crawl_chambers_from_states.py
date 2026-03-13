#!/usr/bin/env python3
"""Read state list from `output/state_chambers.xlsx`, crawl each state page,
extract chamber site names and URLs, and append them as a new sheet in
`output/state_chambers.xlsx`.

Usage:
    python3 scripts/crawl_chambers_from_states.py --input output/state_chambers.xlsx
"""
import argparse
import logging
import time
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup


def is_external_link(url: str) -> bool:
    if not url:
        return False
    url = url.strip()
    if url.startswith("mailto:") or url.startswith("javascript:"):
        return False
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return False
    # treat officialusa links as internal
    if "officialusa.com" in parsed.netloc.lower():
        return False
    return True


def fetch(url: str) -> str:
    tries = 0
    backoff = 1.0
    while tries < 6:
        try:
            r = requests.get(url, timeout=20)
            r.raise_for_status()
            return r.text
        except requests.HTTPError as e:
            status = getattr(e.response, 'status_code', None)
            if status == 429:
                logging.warning("429 received for %s, backing off %ss", url, backoff)
                time.sleep(backoff)
                backoff *= 2
                tries += 1
                continue
            raise
        except requests.RequestException:
            tries += 1
            logging.warning("Request failed for %s, retrying in %ss", url, backoff)
            time.sleep(backoff)
            backoff *= 2
    raise RuntimeError(f"Failed to fetch {url} after retries")


def extract_chamber_links(html: str, base_url: str) -> list:
    soup = BeautifulSoup(html, "html.parser")
    found = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        full = urljoin(base_url, href)
        if not is_external_link(full):
            continue
        if full in seen:
            continue
        seen.add(full)
        text = a.get_text(" ", strip=True) or full
        found.append({"chamber_name": text, "chamber_url": full})
    return found


DEFAULT_WORKBOOK = "output/state_chambers.xlsx"


def crawl_states(input_xlsx: str, output_xlsx: str = DEFAULT_WORKBOOK):
    df_states = pd.read_excel(input_xlsx)
    all_rows = []
    for idx, row in df_states.iterrows():
        state = row.get("state") or ""
        url = row.get("url")
        if not url or not isinstance(url, str):
            continue
        logging.info("Crawling state %s -> %s", state, url)
        try:
            html = fetch(url)
        except Exception as e:
            logging.warning("Failed to fetch %s: %s", url, e)
            continue
        items = extract_chamber_links(html, url)
        for it in items:
            it["state"] = state
            it["source"] = url
            all_rows.append(it)
        # polite pause
        time.sleep(0.4)

    if not all_rows:
        logging.info("No chamber links found")
        return

    df = pd.DataFrame(all_rows)

    # append the new sheet to the existing workbook (replace if exists)
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    try:
        wb = load_workbook(output_xlsx)
        if "chambers" in wb.sheetnames:
            del wb["chambers"]
        ws = wb.create_sheet("chambers")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(output_xlsx)
    except FileNotFoundError:
        # write new file
        with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="chambers", index=False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", default="output/state_chambers.xlsx")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    crawl_states(args.input)


if __name__ == "__main__":
    main()
