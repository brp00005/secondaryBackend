import time
import json
import logging
from typing import List, Dict, Iterable, Optional, Tuple
from urllib.parse import urlparse
from pathlib import Path

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)


class DuckDuckGoJobBoardCrawler:
    """SERP crawler specialized for discovering job board domains.

    The original implementation targeted DuckDuckGo's lightweight HTML
    endpoint, but certain engines (e.g. Brave Search) are far more tolerant of
    automated requests. The class now supports multiple backends which are
    largely interchangeable from the caller's perspective; the only
    difference is the HTML selectors used to extract outbound result links.
    """

    # default parameters for supported engines
    ENGINES = {
        "duckduckgo": {
            "base": "https://html.duckduckgo.com/html/",
            "method": "POST",
            "parse_fn": "_parse_duckduckgo",
        },
        "brave": {
            "base": "https://search.brave.com/search",
            "method": "GET",
            "parse_fn": "_parse_brave",
        },
    }

    # keywords we look for in domains/paths
    JOB_KEYWORDS = [
        "job",
        "jobs",
        "careers",
        "work",
        "hiring",
        "apply",
        "recruit",
        "employment",
    ]

    # phrases we scan for on homepages
    PAGE_KEYWORDS = [
        "search jobs",
        "post a job",
        "job listings",
        "find jobs",
        "apply now",
        "browse jobs",
    ]

    # known job aggregators
    JOB_AGGREGATORS = {
        "indeed.com",
        "linkedin.com",
        "glassdoor.com",
        "ziprecruiter.com",
        "monster.com",
        "careerbuilder.com",
        "snagajob.com",
        "joblist.com",
        "betterteam.com",
        "simplyhired.com",
        "crunchboard.com",
        "dice.com",
        "builtin.com",
        "techcrunch.com",
        "stackoverflow.com",
    }

    # common career page paths to try
    CAREER_PAGE_PATHS = [
        "/careers",
        "/jobs",
        "/work-with-us",
        "/careers/jobs",
        "/join-us",
        "/opportunities",
        "/recruitment",
        "/career",
        "/hiring",
        "/employment",
    ]

    def __init__(
        self,
        rate_limit: float = 1.0,
        user_agent: Optional[str] = None,
        timeout: int = 15,
        engine: str = "duckduckgo",
    ):
        self.rate_limit = rate_limit
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": user_agent or "duckduckgo-jobboard-crawler/1.0 (+https://example.com)",
            "Accept-Encoding": "gzip, deflate"
        })
        self.timeout = timeout
        self.engine = engine.lower()
        if self.engine not in self.ENGINES:
            raise ValueError(f"unsupported engine: {engine}")

    # --- checkpoint/resume ---------------------------------------------------
    @staticmethod
    def load_checkpoint(checkpoint_file: str = ".crawler_checkpoint.json") -> Dict:
        """Load the last checkpoint for resuming crawl."""
        try:
            p = Path(checkpoint_file)
            if p.exists():
                with open(p, "r") as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"could not load checkpoint: {e}")
        return {"last_query_index": -1, "discovered_count": 0}

    @staticmethod
    def save_checkpoint(data: Dict, checkpoint_file: str = ".crawler_checkpoint.json") -> None:
        """Save checkpoint for resuming later."""
        try:
            with open(checkpoint_file, "w") as f:
                json.dump(data, f)
        except Exception as e:
            logger.warning(f"could not save checkpoint: {e}")

    # --- categorization -------------------------------------------------
    @staticmethod
    def is_job_aggregator(url: str) -> bool:
        """Check if URL is a known job aggregator."""
        domain = urlparse(url).netloc.lower()
        domain = domain.replace("www.", "")
        return domain in DuckDuckGoJobBoardCrawler.JOB_AGGREGATORS

    def find_career_page(self, domain: str) -> Optional[str]:
        """Try to find career page for a company domain."""
        for path in self.CAREER_PAGE_PATHS:
            career_url = f"https://{domain}{path}"
            try:
                r = self.session.head(career_url, timeout=5, allow_redirects=True)
                if r.status_code < 400:
                    # check if it's actually a careers page
                    if self._looks_like_career_page(career_url):
                        return career_url
            except requests.RequestException:
                pass
        return None

    def _looks_like_career_page(self, url: str) -> bool:
        """Check if a URL looks like a career/jobs page."""
        try:
            r = self.session.get(url, timeout=5)
            r.raise_for_status()
            text = r.text.lower()
            for kw in self.PAGE_KEYWORDS + ["career", "position", "opening"]:
                if kw in text:
                    return True
        except requests.RequestException:
            pass
        return False

    # --- spreadsheet output --------------------------------------------------
    def save_company_careers_to_excel(self, companies: List[Dict], path: str) -> None:
        """Save company career pages to Excel.
        
        Expected format: [{"domain": "...", "career_page": "...", "title": "..."}]
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Company Careers"
        ws.append(["Domain", "Career Page URL", "Title", "Source"])
        for item in companies:
            ws.append([
                item.get("domain", ""),
                item.get("career_page", ""),
                item.get("title", ""),
                item.get("source", "direct"),
            ])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        wb.save(path)

    def save_aggregators_to_excel(self, aggregators: List[Dict], path: str) -> None:
        """Save job aggregators to Excel.
        
        Expected format: [{"domain": "...", "url": "...", "title": "..."}]
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Aggregators"
        ws.append(["Domain", "URL", "Title"])
        for item in aggregators:
            ws.append([
                item.get("domain", ""),
                item.get("url", ""),
                item.get("title", ""),
            ])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 50
        wb.save(path)

    @staticmethod
    def default_queries() -> List[str]:
        """Return default queries for job board discovery."""
        return [
            "job boards",
            "list of job boards",
            "best job boards 2025",
            "remote job boards",
            "tech job boards",
            "job posting sites",
            "job board directory",
            "job board list",
            "job search sites",
            "job sites",
            "employment websites",
            "career portals",
            "job search engines",
            "US job boards",
            "UK job boards",
            "engineering job boards",
            "healthcare job boards",
        ]

    @staticmethod
    def normalize_url(url: str) -> str:
        """Strip tracking fragments and return canonical form."""
        parsed = urlparse(url)
        scheme = parsed.scheme or "http"
        netloc = parsed.netloc.lower()
        path = parsed.path or "/"
        return f"{scheme}://{netloc}{path}"

    @staticmethod
    def get_domain(url: str) -> str:
        """Extract domain from URL."""
        try:
            return urlparse(url).netloc.lower()
        except Exception:
            return ""

    def is_likely_job_board(self, url: str) -> bool:
        """Check if URL matches job board heuristics."""
        u = self.normalize_url(url)
        domain = self.get_domain(u)
        path = urlparse(u).path.lower()

        for kw in self.JOB_KEYWORDS:
            if kw in domain or kw in path:
                return True
        return False

    def _check_page_for_keywords(self, url: str) -> bool:
        """Fetch homepage and scan for job-related phrases."""
        try:
            r = self.session.get(url, timeout=self.timeout)
            r.raise_for_status()
            text = r.text.lower()
            for kw in self.PAGE_KEYWORDS:
                if kw in text:
                    return True
        except requests.RequestException:
            pass
        return False

    def confirm_job_board(self, url: str, deep: bool = False) -> bool:
        """Confirm whether a URL is a real job board.

        If `deep` is False, this falls back to lightweight heuristics
        (`is_likely_job_board`). If `deep` is True, fetch the page and look
        for stronger signals: JSON-LD JobPosting, job-related headings,
        'apply'/'job' keywords in visible text, or multiple links that look
        like job listings.
        """
        if not url:
            return False

        # quick heuristic first
        if not deep:
            return self.is_likely_job_board(url)

        try:
            r = self.session.get(url, timeout=self.timeout)
            r.raise_for_status()
            text = r.text
        except requests.RequestException:
            return False

        lower = text.lower()

        # strong signal: structured JobPosting JSON-LD
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(text, "html.parser")
            for script in soup.find_all("script", type="application/ld+json"):
                try:
                    import json
                    payload = json.loads(script.string or "{}")
                    # payload may be a list
                    if isinstance(payload, list):
                        items = payload
                    else:
                        items = [payload]
                    for it in items:
                        it_type = it.get("@type") if isinstance(it, dict) else None
                        if it_type and "job" in it_type.lower():
                            return True
                except Exception:
                    continue
        except Exception:
            pass

        # check for job-related keywords in visible text
        for kw in self.PAGE_KEYWORDS + ["job", "jobs", "apply", "position", "opening", "vacancy"]:
            if kw in lower:
                return True

        # check for multiple links that look like job postings
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(text, "html.parser")
            anchors = soup.find_all("a", href=True)
            job_like = 0
            for a in anchors:
                href = a["href"].lower()
                if any(p in href for p in ["/jobs", "/careers", "/positions", "/openings", "job/"]):
                    job_like += 1
                # also check anchor text
                txt = (a.get_text(" ", strip=True) or "").lower()
                if any(k in txt for k in ["apply", "job", "position", "opening"]):
                    job_like += 1
                if job_like >= 3:
                    return True
        except Exception:
            pass

        return False

    # --- career link discovery & verification ------------------------------
    def find_career_links(self, site_url: str) -> List[str]:
        """Find candidate career/job links on a company site homepage.

        Returns resolved absolute URLs to candidate career pages.
        """
        try:
            r = self.session.get(site_url, timeout=self.timeout)
            r.raise_for_status()
            html = r.text
        except requests.RequestException:
            return []

        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        candidates = []
        keywords = ["careers", "jobs", "openings", "positions", "join-us", "work-with-us", "vacancies", "apply"]
        for a in soup.find_all(["a", "button"], href=True):
            text = (a.get_text(" ", strip=True) or "").lower()
            href = a.get("href")
            if not href:
                continue
            if any(k in text for k in keywords) or any(k in href.lower() for k in keywords):
                # resolve relative URL
                from urllib.parse import urljoin
                resolved = urljoin(site_url, href)
                if resolved not in candidates:
                    candidates.append(resolved)

        # also look for buttons without href but with onclicks containing URLs
        for btn in soup.find_all("button"):
            onclick = btn.get("onclick") or ""
            if onclick and any(k in btn.get_text(" ", strip=True).lower() for k in keywords):
                import re
                m = re.search(r"(https?://[^"]+)", onclick)
                if m:
                    url = m.group(1)
                    if url not in candidates:
                        candidates.append(url)

        return candidates

    def verify_jobs_on_page(self, page_url: str) -> bool:
        """Verify that a page contains actual job postings.

        Heuristics:
        - JSON-LD JobPosting present
        - multiple anchors with job-related hrefs/text
        - presence of strings like 'apply now', 'job description', 'posted'
        """
        try:
            r = self.session.get(page_url, timeout=self.timeout)
            r.raise_for_status()
            text = r.text
        except requests.RequestException:
            return False

        lower = text.lower()
        if "jobposting" in lower or "job posting" in lower:
            return True

        # JSON-LD check
        try:
            from bs4 import BeautifulSoup
            import json
            soup = BeautifulSoup(text, "html.parser")
            for script in soup.find_all("script", type="application/ld+json"):
                try:
                    payload = json.loads(script.string or "{}")
                    if isinstance(payload, list):
                        items = payload
                    else:
                        items = [payload]
                    for it in items:
                        t = it.get("@type") if isinstance(it, dict) else None
                        if t and "job" in t.lower():
                            return True
                except Exception:
                    continue
        except Exception:
            pass

        # look for several job-like anchors or keywords
        job_like = 0
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"].lower()
                txt = (a.get_text(" ", strip=True) or "").lower()
                if any(p in href for p in ["/jobs", "/careers", "/positions", "/openings", "job/"]):
                    job_like += 1
                if any(k in txt for k in ["apply", "apply now", "job", "position", "opening", "vacancy"]):
                    job_like += 1
                if job_like >= 3:
                    return True
        except Exception:
            pass

        # keyword sniff
        for phrase in ["apply now", "job description", "posted", "salary"]:
            if phrase in lower:
                return True

        return False

    # --- chamber of commerce directory scraping ---------------------------
    def find_member_directory(self, site_url: str) -> Optional[str]:
        """Find a member directory URL on a Chamber of Commerce site (if present)."""
        try:
            r = self.session.get(site_url, timeout=self.timeout)
            r.raise_for_status()
            html = r.text
        except requests.RequestException:
            return None

        from bs4 import BeautifulSoup
        from urllib.parse import urljoin
        soup = BeautifulSoup(html, "html.parser")
        keywords = ["member", "directory", "members", "member directory", "business directory"]
        for a in soup.find_all("a", href=True):
            txt = (a.get_text(" ", strip=True) or "").lower()
            href = a.get("href")
            if any(k in txt for k in keywords) or any(k in href.lower() for k in keywords):
                return urljoin(site_url, href)
        return None

    def scrape_member_directory(self, directory_url: str, max_pages: int = 50) -> List[Dict]:
        """Scrape a member directory and return list of {'name','url'}.

        Attempts to page through simple alphabetical or pagination schemes.
        """
        members: List[Dict] = []
        try:
            r = self.session.get(directory_url, timeout=self.timeout)
            r.raise_for_status()
            html = r.text
        except requests.RequestException:
            return members

        from bs4 import BeautifulSoup
        from urllib.parse import urljoin
        soup = BeautifulSoup(html, "html.parser")

        # Simple member row extraction heuristics
        # Look for lists / tables of companies
        def extract_from_soup(soup_page):
            found = []
            # common patterns: li items, table rows, divs with class member
            for li in soup_page.find_all("li"):
                a = li.find("a", href=True)
                if a:
                    name = a.get_text(strip=True)
                    href = urljoin(directory_url, a["href"])
                    if name and href:
                        found.append({"name": name, "url": href})
            for tr in soup_page.find_all("tr"):
                a = tr.find("a", href=True)
                if a:
                    name = a.get_text(strip=True)
                    href = urljoin(directory_url, a["href"])
                    if name and href:
                        found.append({"name": name, "url": href})
            # fallback: anchors under divs
            for div in soup_page.find_all("div"):
                a = div.find("a", href=True)
                if a and len(a.get_text(strip=True)) > 2:
                    name = a.get_text(strip=True)
                    href = urljoin(directory_url, a["href"])
                    found.append({"name": name, "url": href})
            return found

        members.extend(extract_from_soup(soup))

        # Try to follow A-Z links if present
        az_links = []
        for a in soup.find_all("a", href=True):
            txt = (a.get_text(" ", strip=True) or "").strip()
            if len(txt) == 1 and txt.isalpha():
                az_links.append(urljoin(directory_url, a["href"]))

        visited = set()
        pages = 0
        for link in az_links:
            if pages >= max_pages:
                break
            if link in visited:
                continue
            visited.add(link)
            try:
                r = self.session.get(link, timeout=self.timeout)
                r.raise_for_status()
                soup2 = BeautifulSoup(r.text, "html.parser")
                members.extend(extract_from_soup(soup2))
            except requests.RequestException:
                continue
            pages += 1

        # Try basic pagination (next links)
        next_link = None
        try:
            next_a = soup.find("a", string=lambda s: s and "next" in s.lower())
            if next_a and next_a.get("href"):
                next_link = urljoin(directory_url, next_a.get("href"))
        except Exception:
            next_link = None

        pages = 0
        while next_link and pages < max_pages:
            if next_link in visited:
                break
            visited.add(next_link)
            try:
                r = self.session.get(next_link, timeout=self.timeout)
                r.raise_for_status()
                soup3 = BeautifulSoup(r.text, "html.parser")
                members.extend(extract_from_soup(soup3))
                na = soup3.find("a", string=lambda s: s and "next" in s.lower())
                next_link = urljoin(directory_url, na.get("href")) if na and na.get("href") else None
            except Exception:
                break
            pages += 1

        # dedupe by url
        seen = set()
        out: List[Dict] = []
        for m in members:
            u = m.get("url")
            if not u:
                continue
            if u in seen:
                continue
            seen.add(u)
            out.append(m)

        return out

    def confirm_company_site(self, url: str, deep: bool = False) -> bool:
        """Heuristic check whether a URL/domain is an actual company website.

        This excludes known forum/blog/social domains (reddit, medium, github,
        twitter, facebook, quora, discourse, etc.) and optionally performs a
        deeper page fetch to look for company signals like 'about', 'contact',
        copyright notices, or legal entity identifiers (Inc, LLC) when
        `deep=True`.
        """
        if not url:
            return False

        # normalize domain
        try:
            from urllib.parse import urlparse
            domain = urlparse(url).netloc.lower()
            domain = domain.replace("www.", "")
        except Exception:
            domain = url.lower()

        # blacklist of common forum/social/blog domains
        FORUM_DOMAINS = [
            "reddit.com",
            "medium.com",
            "github.com",
            "twitter.com",
            "facebook.com",
            "quora.com",
            "discourse.org",
            "forums.",
            "forum.",
        ]
        for bad in FORUM_DOMAINS:
            if bad in domain:
                return False

        # lightweight check passes (domain isn't a known forum)
        if not deep:
            return True

        # deep check: fetch homepage and look for company signals
        try:
            r = self.session.get(f"https://{domain}", timeout=self.timeout)
            r.raise_for_status()
            text = r.text.lower()
        except requests.RequestException:
            # Try http fallback
            try:
                r = self.session.get(f"http://{domain}", timeout=self.timeout)
                r.raise_for_status()
                text = r.text.lower()
            except requests.RequestException:
                return False

        signals = ["about", "contact", "©", "copyright", "inc", "llc", "headquarters", "our team", "leadership", "careers"]
        for s in signals:
            if s in text:
                return True

        return False

    def filter_urls(self, urls: Iterable[str], verify_content: bool = False) -> List[str]:
        """Filter URLs that look like job boards."""
        filtered: List[str] = []
        for u in urls:
            if self.is_likely_job_board(u):
                filtered.append(u)
            elif verify_content and self._check_page_for_keywords(u):
                filtered.append(u)
        return filtered

    def extract_links(self, items: List[Dict]) -> List[str]:
        """Extract URLs from search results."""
        return [self.normalize_url(it["url"]) for it in items if it.get("url")]

    def dedupe(self, urls: Iterable[str]) -> List[str]:
        """Remove duplicate URLs."""
        seen = set()
        out: List[str] = []
        for u in urls:
            if u not in seen:
                seen.add(u)
                out.append(u)
        return out

    def _fetch_search_page(self, query: str, offset: int = 0) -> Optional[str]:
        """Fetch a search results page from the configured engine."""
        cfg = self.ENGINES[self.engine]
        method = cfg["method"].upper()
        params = {}
        data = {}

        if self.engine == "duckduckgo":
            data = {"q": query, "s": str(offset)}
        elif self.engine == "brave":
            params = {"q": query, "source": "web"}
        else:
            params = {"q": query}

        try:
            if method == "POST":
                resp = self.session.post(cfg["base"], data=data, params=params, timeout=self.timeout)
            else:
                resp = self.session.get(cfg["base"], params=params, timeout=self.timeout)
            resp.raise_for_status()
            text = resp.text

            # detect bot challenges
            if self.engine == "duckduckgo" and (
                "Select all squares containing a duck" in text or "error-lite" in text
            ):
                logger.warning("received bot challenge from DuckDuckGo; results suppressed")
                return None
            if self.engine == "brave" and "Sorry your network appears" in text:
                logger.warning("brave search appears to have blocked the request")
                return None
            return text
        except requests.RequestException as e:
            logger.warning("request failed: %s", e)
            return None

    def _parse_results(self, html: str) -> List[Dict]:
        """Parse results using engine-specific parser."""
        fn_name = self.ENGINES[self.engine]["parse_fn"]
        parse_fn = getattr(self, fn_name)
        return parse_fn(html)

    def _parse_duckduckgo(self, html: str) -> List[Dict]:
        """Parse DuckDuckGo HTML results."""
        soup = BeautifulSoup(html, "html.parser")
        results = []
        for div in soup.find_all("div", class_="result"):
            a = div.find("a")
            if not a or not a.get("href"):
                continue
            title = a.get_text(strip=True)
            url = a.get("href")
            snippet_tag = div.find("a", class_="result__snippet") or div.find("div", class_="result__snippet")
            snippet = snippet_tag.get_text(strip=True) if snippet_tag else ""
            results.append({"title": title, "url": url, "snippet": snippet})
        return results

    def _parse_brave(self, html: str) -> List[Dict]:
        """Parse Brave Search HTML results."""
        soup = BeautifulSoup(html, "html.parser")
        results = []
        for div in soup.find_all("div", class_="result-content"):
            a = div.find("a", href=True)
            if not a:
                continue
            href = a["href"]
            title = a.get_text(strip=True)
            results.append({"title": title, "url": href, "snippet": ""})
        return results

    def search(self, query: str, pages: int = 1, results_per_page: int = 50) -> List[Dict]:
        """Search for a query and return all results."""
        all_results: List[Dict] = []
        for p in range(pages):
            offset = p * results_per_page
            html = self._fetch_search_page(query, offset=offset)
            if not html:
                break
            page_results = self._parse_results(html)
            if not page_results:
                break
            all_results.extend(page_results)
            time.sleep(self.rate_limit)
        return all_results

    def save(self, items: List[Dict], path: str) -> None:
        """Save results to JSON file."""
        with open(path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)

    def save_domains_to_excel(self, domains: List[str], path: str) -> None:
        """Save domains to Excel file."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Boards"
        ws.append(["Domain", "URL"])
        for domain in domains:
            ws.append([domain, f"https://{domain}"])
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        wb.save(path)

    def save_results_to_excel(self, items: List[Dict], path: str) -> None:
        """Save search results to Excel file."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Title", "URL", "Query"])
        for item in items:
            ws.append([item.get("title", ""), item.get("url", ""), item.get("query", "")])
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30
        wb.save(path)


if __name__ == "__main__":
    crawler = DuckDuckGoJobBoardCrawler(engine="brave")
    queries = DuckDuckGoJobBoardCrawler.default_queries()
    print(f"running {len(queries)} queries with Brave Search")
    all_items: List[Dict] = []
    for q in queries:
        res = crawler.search(q, pages=1)
        for r in res:
            r.setdefault("query", q)
        all_items.extend(res)
    urls = crawler.extract_links(all_items)
    urls = crawler.dedupe(urls)
    print(f"extracted {len(urls)} unique urls")

    # apply simple heuristics
    boards = crawler.filter_urls(urls, verify_content=False)
    print(f"{len(boards)} candidates after heuristic filtering")

    crawler.save([{"url": u} for u in boards], "candidates.json")
    print("saved candidate urls to candidates.json")
